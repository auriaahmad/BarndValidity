from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
from typing import List, Tuple

def generate_qr_excel(qr_codes: List[Tuple[str, BytesIO, str]], school_name: str, batch_number: str) -> BytesIO:
    """
    Generate an Excel file containing QR codes and their information.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "QR Codes"
    
    # Add headers
    headers = ["School Name", "Batch Number", "Product Code", "Verification URL"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Add QR codes and data
    for row, (unique_code, qr_image, verification_url) in enumerate(qr_codes, 2):
        # Add data
        ws.cell(row=row, column=1, value=school_name)
        ws.cell(row=row, column=2, value=batch_number)
        ws.cell(row=row, column=3, value=unique_code)
        ws.cell(row=row, column=4, value=verification_url)
        
        # Add QR code image
        qr_image.seek(0)
        img = XLImage(qr_image)
        img.width = 100
        img.height = 100
        ws.add_image(img, f'E{row}')
    
    # Adjust column widths
    for col in range(1, 5):
        ws.column_dimensions[chr(64 + col)].width = 20
    
    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer 